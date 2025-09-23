import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
from openpyxl.styles import Border, Side, PatternFill

# Estrutura de dados em memória
foranias = {}  # { 'Forania 1': { 'Grupo 1': [ {servo}, ... ] } }

# Funções
def adicionar_forania():
    nome = entrada_forania.get().strip()
    if nome:
        if nome not in foranias:
            foranias[nome] = {}
            atualizar_combo_forania()
            renderizar()
            entrada_forania.delete(0, tk.END)
        else:
            messagebox.showwarning("Aviso", "Forania já existe.")

def adicionar_grupo():
    forania = combo_forania.get()
    nome = entrada_grupo.get().strip()
    if forania and nome:
        if nome not in foranias[forania]:
            foranias[forania][nome] = []
            atualizar_combo_grupo()
            renderizar()
            entrada_grupo.delete(0, tk.END)
        else:
            messagebox.showwarning("Aviso", "Grupo já existe.")

def adicionar_servo():
    forania = combo_forania.get()
    grupo = combo_grupo.get()
    nome = entrada_servo_nome.get().strip()
    if not (forania and grupo and nome):
        messagebox.showwarning("Aviso", "Preencha todos os campos.")
        return

    servo = {
        "nome": nome,
        "modulo_basico": var_modulo.get(),
        "exp_oracao": var_oracao.get(),
        "apostila1": var_ap1.get(),
        "apostila2": var_ap2.get()
    }

    foranias[forania][grupo].append(servo)
    entrada_servo_nome.delete(0, tk.END)
    for var in [var_modulo, var_oracao, var_ap1, var_ap2]:
        var.set(False)
    renderizar()

def atualizar_combo_forania():
    combo_forania['values'] = list(foranias.keys())

def atualizar_combo_grupo(*args):
    forania = combo_forania.get()
    if forania:
        combo_grupo['values'] = list(foranias[forania].keys())

def exportar_excel():
    if not foranias:
        messagebox.showwarning("Aviso", "Nenhum dado para exportar.")
        return

    caminho = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if caminho:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Servos"

        thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
        )

        row_index = 1  # controla a linha atual

        for forania, grupos in foranias.items():
            start_row = row_index 

            # Cabeçalho da Forania
            #ws.cell(row=row_index, column=1, value="").font = openpyxl.styles.Font(bold=True)
            # ws.cell(row=row_index, column=1, value=forania).font = openpyxl.styles.Font(bold=True)
            # cell_forania = ws.cell(row=row_index, column=1)
            # cell_forania.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

            fill_forania = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

            for col in range(1, 7):  # da coluna 1 até 6
                cell = ws.cell(row=row_index, column=col)
                if col == 1:  # escreve o nome da Forania só na primeira coluna
                    cell.value = forania
                    cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = fill_forania

            row_index += 2

            for grupo, servos in grupos.items():
                # Cabeçalho do Grupo
                #ws.cell(row=row_index, column=1, value="").font = openpyxl.styles.Font(bold=True)
                # ws.cell(row=row_index, column=1, value=grupo).font = openpyxl.styles.Font(bold=True)

                
             fill_grupo = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

            for col in range(1, 7):  # da coluna 1 até 6
                    cell = ws.cell(row=row_index, column=col)
                    if col == 1:  # escreve o nome do Grupo só na primeira coluna
                        cell.value = grupo
                        cell.font = openpyxl.styles.Font(bold=True)
                    cell.fill = fill_grupo
            row_index += 1

            # Cabeçalhos da tabela
            ws.cell(row=row_index, column=2, value="Nome")
            ws.cell(row=row_index, column=3, value="Módulo Básico")
            ws.cell(row=row_index, column=4, value="Exp. Oração")
            ws.cell(row=row_index, column=5, value="Apostila 1")
            ws.cell(row=row_index, column=6, value="Apostila 2")
            row_index += 1

                # Dados dos servos
            for servo in servos:
                ws.cell(row=row_index, column=2, value=servo["nome"])
                ws.cell(row=row_index, column=3, value="Sim" if servo["modulo_basico"] else "Não")
                ws.cell(row=row_index, column=4, value="Sim" if servo["exp_oracao"] else "Não")
                ws.cell(row=row_index, column=5, value="Sim" if servo["apostila1"] else "Não")
                ws.cell(row=row_index, column=6, value="Sim" if servo["apostila2"] else "Não")
                row_index += 1

                # Linha em branco entre grupos
                row_index += 1
            
            end_row = row_index - 1

            # Aplicar bordas
            for r in range(start_row, end_row + 1):
                for c in range(1, 7):
                    ws.cell(row=r, column=c).border = thin_border

            # Linha em branco entre foranias
            row_index += 1

        wb.save(caminho)
        messagebox.showinfo("Sucesso", f"Dados exportados com sucesso para:\n{caminho}")


def renderizar():
    for widget in frame_resultados.winfo_children():
        widget.destroy()

    for forania, grupos in foranias.items():
        lf_forania = ttk.LabelFrame(frame_resultados, text=f"Forania: {forania}")
        lf_forania.pack(fill='x', padx=5, pady=5, anchor='w')

        for grupo, servos in grupos.items():
            lf_grupo = ttk.LabelFrame(lf_forania, text=f"Grupo: {grupo}")
            lf_grupo.pack(fill='x', padx=10, pady=5)

            tree = ttk.Treeview(lf_grupo, columns=("Nome", "Módulo Básico", "Exp. Oração", "Apostila 1", "Apostila 2"), show="headings")
            for col in tree["columns"]:
                tree.heading(col, text=col)
                tree.column(col, width=120)
            for servo in servos:
                tree.insert('', 'end', values=(
                    servo["nome"],
                    "Sim" if servo["modulo_basico"] else "Não",
                    "Sim" if servo["exp_oracao"] else "Não",
                    "Sim" if servo["apostila1"] else "Não",
                    "Sim" if servo["apostila2"] else "Não"
                ))
            tree.pack(fill='x')

# UI principal
janela = tk.Tk()
janela.title("Cadastro de Foranias, Grupos e Servos")
janela.geometry("850x700")

# --- Cadastro ---
frame_cadastro = ttk.LabelFrame(janela, text="Cadastro")
frame_cadastro.pack(fill='x', padx=10, pady=10)

# Forania
ttk.Label(frame_cadastro, text="Nova Forania:").grid(row=0, column=0, sticky='e')
entrada_forania = ttk.Entry(frame_cadastro)
entrada_forania.grid(row=0, column=1)
ttk.Button(frame_cadastro, text="Adicionar Forania", command=adicionar_forania).grid(row=0, column=2, padx=5)

# Grupo
ttk.Label(frame_cadastro, text="Nova Grupo:").grid(row=1, column=0, sticky='e')
combo_forania = ttk.Combobox(frame_cadastro, state="readonly")
combo_forania.grid(row=1, column=1)
combo_forania.bind("<<ComboboxSelected>>", atualizar_combo_grupo)

entrada_grupo = ttk.Entry(frame_cadastro)
entrada_grupo.grid(row=1, column=2)
ttk.Button(frame_cadastro, text="Adicionar Grupo", command=adicionar_grupo).grid(row=1, column=3)

# Servo
ttk.Label(frame_cadastro, text="Grupo (Seleção):").grid(row=2, column=0, sticky='e')
combo_grupo = ttk.Combobox(frame_cadastro, state="readonly")
combo_grupo.grid(row=2, column=1)

ttk.Label(frame_cadastro, text="Nome do Servo:").grid(row=3, column=0, sticky='e')
entrada_servo_nome = ttk.Entry(frame_cadastro)
entrada_servo_nome.grid(row=3, column=1)

# Checkbuttons
var_modulo = tk.BooleanVar()
var_oracao = tk.BooleanVar()
var_ap1 = tk.BooleanVar()
var_ap2 = tk.BooleanVar()

ttk.Checkbutton(frame_cadastro, text="Módulo Básico", variable=var_modulo).grid(row=4, column=0, sticky='w')
ttk.Checkbutton(frame_cadastro, text="Exp. Oração", variable=var_oracao).grid(row=4, column=1, sticky='w')
ttk.Checkbutton(frame_cadastro, text="Apostila 1", variable=var_ap1).grid(row=4, column=2, sticky='w')
ttk.Checkbutton(frame_cadastro, text="Apostila 2", variable=var_ap2).grid(row=4, column=3, sticky='w')

ttk.Button(frame_cadastro, text="Adicionar Servo", command=adicionar_servo).grid(row=5, column=1, pady=5)

# --- Exportar ---
ttk.Button(janela, text="Exportar para Excel", command=exportar_excel).pack(pady=5)

# --- Frame com Scrollbar para resultados ---
frame_scroll = ttk.Frame(janela)
frame_scroll.pack(fill='both', expand=True, padx=10, pady=10)

canvas = tk.Canvas(frame_scroll)
scrollbar = ttk.Scrollbar(frame_scroll, orient='vertical', command=canvas.yview)
scrollable_frame = ttk.Frame(canvas)

scrollable_frame.bind(
    "<Configure>",
    lambda e: canvas.configure(
        scrollregion=canvas.bbox("all")
    )
)

canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
canvas.configure(yscrollcommand=scrollbar.set)

canvas.pack(side="left", fill="both", expand=True)
scrollbar.pack(side="right", fill="y")

# Substitui o antigo frame_resultados
frame_resultados = scrollable_frame



janela.mainloop()